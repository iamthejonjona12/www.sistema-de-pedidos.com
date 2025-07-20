#quiero que tenga funcion donde diga total ganado en 1dia y asi que  para todo que salga todo lo ganado en el excel
import flet as ft  
from ftplib import FTP
import openpyxl
import pandas as pd
from datetime import datetime
import os
from escpos.printer import Usb

def imprimir_recibo(cliente, pedidos):
    try:
        # ⚠️ Reemplaza los valores por los de tu impresora
        p = Usb(0x04b8, 0x0e15)  # (vendor_id, product_id)
        p.text("POLLERÍA QUEENCY\n")
        p.text(f"Cliente: {cliente}\n")
        p.text(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        p.text("-" * 32 + "\n")

        total = 0
        for i, item in enumerate(pedidos, 1):
            p.text(f"{i}. {item['nombre'][:20]}\n")
            p.text(f"    Bs {item['precio']:.2f}\n")
            total += item["precio"]

        p.text("-" * 32 + "\n")
        p.text(f"TOTAL: Bs {total:.2f}\n")
        p.text("\nGRACIAS POR SU COMPRA\n\n\n")
        p.cut()
    except Exception as ex:
        print("Error al imprimir:", ex)


excel_file = "pedidos_polleria.xlsx"

def main(page: ft.Page):
    page.title = "Pollería de queency"
    page.bgcolor = "#AA4810"  # Rojo fuerte
    page.window_height = 950
    page.scroll = "auto"

    pedidos = []
    pedido_id = 1
    platos_menu = []
    refrescos_menu = []

    # 🆕 Campos
    nombre_cliente = ft.TextField(label="Nombre completo del cliente", width=409, color="#50A137")

    # 🆕 Nuevos campos para acompañamiento y parte del pollo
    acompanamiento_dropdown = ft.Dropdown(
        label="Acompañamiento",
        options=[
            ft.dropdown.Option("Pura papa"),
            ft.dropdown.Option("Arroz y papa"),
            ft.dropdown.Option("Fideo"),
              # Opción vacía para no seleccionar nada
        ],  # Valor por defecto vacío  
        width=205,
    )

    parte_pollo_dropdown = ft.Dropdown(
        label="Parte del pollo",
        options=[
            ft.dropdown.Option("Pierna"),
            ft.dropdown.Option("Pecho"),
            ft.dropdown.Option("Ala"),
            ft.dropdown.Option("Entre pierna")
        ],
        width=195
    )

    # CRUD Helpers
    nuevo_plato = ft.TextField(label="Nombre del plato")
    precio_plato = ft.TextField(label="Precio", width=100)
    nuevo_refresco = ft.TextField(label="Nombre del refresco")
    precio_refresco = ft.TextField(label="Precio", width=100)

    def actualizar_checkbox():
        checkboxes_platos.controls.clear()
        checkboxes_refrescos.controls.clear()
        for p in platos_menu:
            checkboxes_platos.controls.append(ft.Checkbox(label=p["nombre"], data=p["precio"]))
        for r in refrescos_menu:
            checkboxes_refrescos.controls.append(ft.Checkbox(label=r["nombre"], data=r["precio"]))
        page.update()

    def agregar_plato(e):
        if nuevo_plato.value and precio_plato.value:
            import re
            match = re.search(r"\d+(\.\d+)?", precio_plato.value)
            precio = float(match.group()) if match else 0.0
            platos_menu.append({"nombre": nuevo_plato.value, "precio": precio})
            nuevo_plato.value = ""
            precio_plato.value = ""
            actualizar_checkbox()

    def agregar_refresco(e):
        if nuevo_refresco.value and precio_refresco.value:
            import re
            match = re.search(r"\d+(\.\d+)?", precio_refresco.value)
            precio = float(match.group()) if match else 0.0
            refrescos_menu.append({"nombre": nuevo_refresco.value, "precio": precio})
            nuevo_refresco.value = ""
            precio_refresco.value = ""
            actualizar_checkbox()

    def eliminar_refresco(e):
        refrescos_menu[:] = [r for r in refrescos_menu if r["nombre"] != nuevo_refresco.value]
        actualizar_checkbox()

    def eliminar_plato(e):
        platos_menu[:] = [p for p in platos_menu if p["nombre"] != nuevo_plato.value]
        actualizar_checkbox()

    def realizar_pedido(e):
        nonlocal pedido_id
        if not nombre_cliente.value.strip():
            page.dialog = ft.AlertDialog(title=ft.Text("Ingrese nombre del cliente"))
            page.dialog.open = True
            page.update()
            return

        for cb in checkboxes_platos.controls:
            if cb.value:
                nombre = cb.label
                if acompanamiento_dropdown.value:
                    nombre += f" + {acompanamiento_dropdown.value}"
                if parte_pollo_dropdown.value:
                    nombre += f" ({parte_pollo_dropdown.value})"
                pedidos.append({"id": pedido_id, "nombre": nombre, "precio": float(cb.data)})
                pedido_id += 1

        for cb in checkboxes_refrescos.controls:
            if cb.value:
                pedidos.append({"id": pedido_id, "nombre": cb.label, "precio": float(cb.data)})
                pedido_id += 1

        actualizar_lista()

    total_text = ft.Text(value="Total del pedido: Bs 0")

    def actualizar_lista():
        lista_pedidos.controls.clear()
        total = 0

        # 🔁 Función para eliminar un pedido
        def eliminar_item(index): 
            pedidos.pop(index)
            actualizar_lista()
  

        for i, pedido in enumerate(pedidos):
            total += pedido["precio"]
            fila = ft.Row([
                ft.Text(f"{i+1}. {pedido['nombre']} - Bs {pedido['precio']}"),
                ft.IconButton(
                    icon=ft.Icons.DELETE,
                    tooltip="Eliminar",
                    icon_color="red",
                    on_click=lambda e, idx=i: eliminar_item(idx)
                )
            ])
            lista_pedidos.controls.append(fila)

        total_text.value = f"Total del pedido: Bs {total}"
        page.update()
        page.update()



    def guardar_excel(e):
        if not pedidos:
            return
        df = pd.DataFrame(pedidos)
        df["cliente"] = nombre_cliente.value
        df["fecha"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if os.path.exists(excel_file):
                df_existente = pd.read_excel(excel_file)
                dfs_to_concat = [df_existente, df]
                # Remove empty or all-NA DataFrames
                dfs_to_concat = [d for d in dfs_to_concat if not d.empty and not d.isna().all(axis=None)]
                if dfs_to_concat:
                    df = pd.concat(dfs_to_concat, ignore_index=True)
            df.to_excel(excel_file, index=False)
            page.snack_bar = ft.SnackBar(ft.Text("Guardado en Excel"))
            imprimir_recibo(nombre_cliente.value, pedidos)
            page.snack_bar.open = True
            page.update()
        except PermissionError:
            page.dialog = ft.AlertDialog(title=ft.Text("Error: El archivo Excel está abierto. Por favor ciérrelo e intente de nuevo."))
            page.dialog.open = True
            page.update()


    def abrir_excel(e):
        if os.path.exists(excel_file):
            os.system(f'start excel "{excel_file}"')

    def abrir_reporte_ganancias(e):
      if not os.path.exists(excel_file):
        page.dialog = ft.AlertDialog(title=ft.Text("⚠️ El archivo de Excel no existe aún."))
        page.dialog.open = True
        page.update()
        return

    try:
        df = pd.read_excel(excel_file, sheet_name=None)

        if "reporte_ganancias" in df:
            reporte = df["reporte_ganancias"]
            if not reporte.empty:
                ultima = reporte.iloc[-1]
                mensaje = (
                    f"📆 Último Reporte:\n"
                    f"Fecha: {ultima['fecha_reporte']}\n"
                    f"Periodo: {ultima['periodo']}\n"
                    f"Total: Bs {ultima['total_ganado']}"
                )
            else:
                mensaje = "📂 No hay datos en el reporte de ganancias."
        else:
            mensaje = "❌ No se encontró la hoja 'reporte_ganancias'."

    except Exception as ex:
        mensaje = f"❗ Error al abrir el reporte: {ex}"

    page.dialog = ft.AlertDialog(title=ft.Text(mensaje))
    page.dialog.open = True
    page.update()

 
    # 🆕 Funciones para eliminar pedidos y reporte de ganancias
    def eliminar_pedidos_excel(e):
        import openpyxl
        if not os.path.exists(excel_file):
            return
        try:
            wb = openpyxl.load_workbook(excel_file)
            if "Sheet1" in wb.sheetnames:
                std = wb["Sheet1"]
                wb.remove(std)
            nueva = wb.create_sheet("Sheet1")
            nueva.append(["id", "nombre", "precio", "cliente", "fecha"])
            wb.save(excel_file)
            wb.close()
            page.snack_bar = ft.SnackBar(ft.Text("✅ Pedidos eliminados correctamente"))
            page.snack_bar.open = True
            page.update()
        except Exception as ex:
            page.dialog = ft.AlertDialog(title=ft.Text(f"Error: {ex}"))
            page.dialog.open = True
            page.update()


    def eliminar_reporte_ganancias(e):
      import openpyxl
      # Eliminar la hoja de reporte de ganancias
      if not os.path.exists(excel_file):
        return
    try:
        wb = openpyxl.load_workbook(excel_file)
        if "reporte_ganancias" in wb.sheetnames:
            del wb["reporte_ganancias"]
            wb.save(excel_file)
            page.snack_bar = ft.SnackBar(ft.Text("✅ Reporte de ganancias eliminado correctamente"))
            page.snack_bar.open = True
            page.update()
        else:
            page.dialog = ft.AlertDialog(title=ft.Text("No se encontró el reporte de ganancias."))
            page.dialog.open = True
            page.update()
    except Exception as ex:
        page.dialog = ft.AlertDialog(title=ft.Text(f"Error: {ex}"))
        page.dialog.open = True
        page.update()
      

        

    def calcular_ganancias(periodo):
        if not os.path.exists(excel_file):
            return "No hay registros."
        df = pd.read_excel(excel_file)
        if "fecha" not in df.columns or "precio" not in df.columns:
            return "Datos incompletos."
        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        df = df.dropna(subset=["fecha"])
        hoy = datetime.now()
        if periodo == "día":
            inicio = hoy.replace(hour=0, minute=0, second=0)
        elif periodo == "semana":
            inicio = hoy - pd.Timedelta(days=hoy.weekday())
        elif periodo == "mes":
            inicio = hoy.replace(day=1)
        elif periodo == "3meses":
            inicio = hoy - pd.DateOffset(months=3)
        elif periodo == "6meses":
            inicio = hoy - pd.DateOffset(months=6)
        elif periodo == "1año":
            inicio = hoy - pd.DateOffset(years=1)
        else:
            return "Periodo inválido."

        df_filtrado = df[df["fecha"] >= inicio]
        total = df_filtrado["precio"].sum()

        return f"Ganancia total del {periodo}: Bs {total:.2f}"
    def ganancia_total_general():
        if not os.path.exists(excel_file):
            return "No hay registros."
        df = pd.read_excel(excel_file)
        if "precio" not in df.columns:
            return "Datos incompletos."
        total = df["precio"].sum()
        return f"Ganancia total general: Bs {total:.2f}"
    # 🆕 Función para mostrar ganancias
    def mostrar_ganancias(e):
        periodo = filtro_ganancias.value
        resultado = calcular_ganancias(periodo)
        ganancias_texto.value = resultado if resultado is not None else "No hay registros."
        page.update()
        if not os.path.exists(excel_file):
            return
        # Solo procesar si resultado es válido y contiene "Bs"
        if resultado and isinstance(resultado, str) and "Bs" in resultado:
            try:
                total = resultado.split("Bs")[-1].strip()
                registro = pd.DataFrame([{
                    "fecha_reporte": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "periodo": periodo,
                    "total_ganado": float(total)
                }])
                with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    try:
                        existing = pd.read_excel(excel_file, sheet_name="reporte_ganancias")
                        registro = pd.concat([existing, registro], ignore_index=True)
                    except:
                        pass
                    registro.to_excel(writer, sheet_name="reporte_ganancias", index=False)
            except Exception as ex:
                page.dialog = ft.AlertDialog(title=ft.Text(f"Error al guardar reporte: {ex}"))
                page.dialog.open = True
                page.update()
        else:
            # No guardar si resultado no es válido
            pass
        # Manejo de PermissionError
        # (ya está manejado en el bloque try-except)
    
    def mostrar_ganancias_total(e):
        resultado = ganancia_total_general()
        page.dialog = ft.AlertDialog(title=ft.Text(resultado))
        page.dialog.open = True
        page.update()
    # 🆕 Dropdown para seleccionar el periodo de ganancias

    filtro_ganancias = ft.Dropdown(
        label="📊 Reporte de ganancias",
        options=[
            ft.dropdown.Option("día"),
            ft.dropdown.Option("semana"),
            ft.dropdown.Option("mes"),
            ft.dropdown.Option("3meses"),
            ft.dropdown.Option("6meses"),
            ft.dropdown.Option("1año")
           
        ],
        width=200,
        on_change=mostrar_ganancias
    )

    ganancias_texto = ft.Text(value="Ganancias: ")

    checkboxes_platos = ft.Column()
    checkboxes_refrescos = ft.Column()
    total_text = ft.Text(value="Total del pedido: Bs 0")
    lista_pedidos = ft.Column()  # Define lista_pedidos as an ft.Column

    platos_menu.extend([
        {"nombre": "Pollo frito simple 19 BS", "precio": 19 },
        {"nombre": "Pollo frito doble 33 BS", "precio": 33},
        {"nombre": "Pollo a la canasta 18 BS", "precio": 18},
        {"nombre": "Pollo ahumado 16 BS", "precio": 16},
        {"nombre": "Pipocas + refresco 28 BS", "precio": 28},
    ])
    
    ft.Text("Menú de Pollos", size=40, weight="bold")
    ft.Text("Menú de Refrescos", size=40, weight="bold")

    refrescos_menu.extend([
        {"nombre": "Mini Coca Cola 3 BS", "precio": 3},
        {"nombre": "Popular 8 BS", "precio": 8},
        {"nombre": "Coca Cola 1/2 l 12 BS", "precio": 12},
        {"nombre": "Coca Cola 2lT. 18", "precio": 18},
        {"nombre": "Coca Cola 3l 23 BS", "precio": 23},
        {"nombre": "Jugo del Valle 20 BS", "precio": 20},
        {"nombre": "Mini Jugo del Valle 7 BS", "precio": 7},
    ])
    actualizar_checkbox()

    page.add(
        ft.Stack([
            ft.Image(src="icon.png", fit=ft.ImageFit.COVER, width=1600, height=1500),
            ft.Column([
                ft.Text("POLLERÍA - QUEENCY", size=38, weight="bold", color="#F8A436"),
                nombre_cliente,
                ft.Row([acompanamiento_dropdown, parte_pollo_dropdown]),  # 🆕
                ft.Divider(),

                ft.Text("Menú de Platos", size=37, weight="bold", color="#6062C4"),
                checkboxes_platos,
                ft.Text("Agregar / Eliminar Plato"),
                ft.Row([nuevo_plato, precio_plato]),
                ft.Row([
                    ft.ElevatedButton("Agregar Plato", on_click=agregar_plato),
                    ft.ElevatedButton("Eliminar Plato", on_click=eliminar_plato, color=ft.Colors.RED)
                ]),

                ft.Divider(),
                ft.Text("Menú de Refresco", size=37, weight="bold", color="#6BDB2A"),
                checkboxes_refrescos,
                ft.Text("Agregar / Eliminar Refresco"),
                ft.Row([nuevo_refresco, precio_refresco]),
                ft.Row([
                    ft.ElevatedButton("Agregar Refresco", on_click=agregar_refresco),
                    ft.ElevatedButton("Eliminar Refresco", on_click=eliminar_refresco, color=ft.Colors.RED)
                ]),

                ft.Divider(),
                ft.ElevatedButton("Realizar Pedido", on_click=realizar_pedido, width=300, height=50, bgcolor="#46AC18", color="black"),
                ft.Text("Pedidos Realizados:", size=30, weight="bold"),
                lista_pedidos,
                total_text,

                ft.Row([
                    ft.ElevatedButton("Guardar en Excel", on_click=guardar_excel),
                    ft.ElevatedButton("Abrir Excel", on_click=abrir_excel),
                    ft.ElevatedButton("Abrir Reporte de Ganancias", on_click=abrir_reporte_ganancias),
                    ft.ElevatedButton("Eliminar Pedidos en Excel", on_click=eliminar_pedidos_excel, color=ft.Colors.RED),
                    ft.ElevatedButton("Eliminar Reporte de Ganancias", on_click=eliminar_reporte_ganancias, color=ft.Colors.RED),
                    ft.ElevatedButton("🖨️ Imprimir Recibo", on_click=lambda e: imprimir_recibo(nombre_cliente.value, pedidos)),
                    ft.ElevatedButton("💵 Total ganado en todo el tiempo", on_click=mostrar_ganancias_total)
                ]),

                ft.Divider(),
                ft.Row([filtro_ganancias, ganancias_texto])
            ])
        ])
    )

ft.app(target=main)
# This code is a simple Flet application for managing a poultry restaurant's orders.
# It allows users to add and remove dishes and drinks, manage customer orders, and print receipts           
